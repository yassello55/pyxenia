#!/usr/bin/env python3
# args:
#   1: input_file (file) - Excel file to clean
"""
Advanced Sales Data Cleaner
Usage:
    python clean_sales_data.py <input.xlsx>

Example:
    python clean_sales_data.py sales_data_messy.xlsx
"""

import sys
import re
import warnings
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore")

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


# ── Config ──────────────────────────────────────────────────────────────────

VALID_STATUSES   = {"Completed", "Pending", "Cancelled", "Refunded", "Processing"}
VALID_REGIONS    = {"North", "South", "East", "West", "Central"}
VALID_CATEGORIES = {"Electronics", "Accessories", "Furniture", "Supplies"}
DATE_FORMATS     = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
    "%B %d, %Y", "%b %d, %Y", "%d-%m-%Y",
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def log(msg: str):
    print(f"  {msg}")


def parse_date(val) -> str | None:
    """Try multiple date formats, return ISO string or None."""
    if pd.isna(val) or str(val).strip() == "":
        return None
    s = str(val).strip()
    # Already a pandas Timestamp
    if isinstance(val, (pd.Timestamp,)):
        return val.strftime("%Y-%m-%d")
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def clean_price(val) -> float | None:
    """Strip currency symbols and convert to float."""
    if pd.isna(val):
        return None
    s = str(val).strip().replace("$", "").replace(",", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def is_valid_email(email: str) -> bool:
    pattern = r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
    return bool(re.match(pattern, email.strip()))


def normalize_phone(phone: str) -> str:
    """Normalize phone to XXX-XXX-XXXX format."""
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return phone  # Return as-is if can't normalize


def title_case_name(name: str) -> str:
    return " ".join(word.capitalize() for word in name.strip().split())


# ── Main Cleaner ─────────────────────────────────────────────────────────────

def clean_sales_data(input_path: str, output_path: str):
    print("\n" + "=" * 60)
    print("  🧹  ADVANCED SALES DATA CLEANER")
    print("=" * 60)
    print(f"\n  Input  : {input_path}")
    print(f"  Output : {output_path}\n")

    df = pd.read_excel(input_path, dtype=str)
    df.columns = df.columns.str.strip()
    original_count = len(df)
    issues = []

    print(f"  📋  Loaded {original_count} rows, {len(df.columns)} columns\n")
    print("─" * 60)

    # ── 1. Remove fully empty rows ───────────────────────────────────────────
    before = len(df)
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    removed = before - len(df)
    if removed:
        log(f"[Empty Rows]     Removed {removed} fully empty row(s)")
        issues.append(f"Removed {removed} empty row(s)")

    # ── 2. Remove duplicates ─────────────────────────────────────────────────
    before = len(df)
    df.drop_duplicates(inplace=True)
    df.reset_index(drop=True, inplace=True)
    dupes = before - len(df)
    if dupes:
        log(f"[Duplicates]     Removed {dupes} duplicate row(s)")
        issues.append(f"Removed {dupes} duplicate row(s)")

    # ── 3. Trim whitespace from all string columns ───────────────────────────
    str_cols = df.select_dtypes(include="object").columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())
    log("[Whitespace]     Stripped leading/trailing spaces from all columns")

    # ── 4. Normalize Order IDs ───────────────────────────────────────────────
    missing_ids = df["Order ID"].isna() | (df["Order ID"] == "")
    if missing_ids.any():
        count = missing_ids.sum()
        # Generate placeholder IDs
        max_existing = df["Order ID"].str.extract(r"(\d+)")[0].dropna().astype(int)
        next_num = max_existing.max() + 1 if not max_existing.empty else 100
        new_ids = [f"ORD-{str(next_num + i).zfill(3)}" for i in range(count)]
        df.loc[missing_ids, "Order ID"] = new_ids
        log(f"[Order ID]       Filled {count} missing Order ID(s) with generated IDs")
        issues.append(f"Generated {count} missing Order ID(s)")

    # ── 5. Standardize dates ─────────────────────────────────────────────────
    bad_dates = 0
    for i, val in df["Order Date"].items():
        parsed = parse_date(val)
        if parsed is None:
            bad_dates += 1
            issues.append(f"Row {i+2}: Could not parse date '{val}'")
            df.at[i, "Order Date"] = "INVALID DATE"
        elif str(val) != parsed:
            df.at[i, "Order Date"] = parsed
    log(f"[Dates]          Standardized to ISO format (YYYY-MM-DD) — {bad_dates} unparseable date(s) flagged")

    # ── 6. Normalize customer names ──────────────────────────────────────────
    fixed_names = 0
    for i, val in df["Customer Name"].items():
        if pd.isna(val) or val.strip() == "":
            continue
        normalized = title_case_name(val)
        if normalized != val:
            df.at[i, "Customer Name"] = normalized
            fixed_names += 1
    if fixed_names:
        log(f"[Names]          Fixed casing/spacing on {fixed_names} customer name(s)")
        issues.append(f"Fixed {fixed_names} customer name(s)")

    # ── 7. Clean & validate emails ───────────────────────────────────────────
    invalid_emails = 0
    for i, val in df["Email"].items():
        if pd.isna(val) or val == "":
            issues.append(f"Row {i+2}: Missing email for '{df.at[i, 'Customer Name']}'")
            continue
        cleaned = val.strip().lower()
        df.at[i, "Email"] = cleaned
        if not is_valid_email(cleaned):
            df.at[i, "Email"] = f"INVALID: {cleaned}"
            invalid_emails += 1
            issues.append(f"Row {i+2}: Invalid email '{cleaned}'")
    log(f"[Emails]         Lowercased all emails — {invalid_emails} invalid email(s) flagged")

    # ── 8. Normalize phone numbers ───────────────────────────────────────────
    for i, val in df["Phone"].items():
        if pd.notna(val) and val.strip():
            df.at[i, "Phone"] = normalize_phone(val)
    log("[Phones]         Normalized phone numbers to XXX-XXX-XXXX format")

    # ── 9. Clean numeric columns ─────────────────────────────────────────────
    for col in ["Unit Price", "Total Price"]:
        fixed = 0
        cleaned_vals = []
        for i, val in df[col].items():
            cleaned = clean_price(val)
            if cleaned is None:
                cleaned_vals.append(None)
            else:
                if str(val).strip() != str(cleaned):
                    fixed += 1
                cleaned_vals.append(str(cleaned))
        df[col] = cleaned_vals
        if fixed:
            log(f"[{col:<12}] Removed currency symbols from {fixed} cell(s)")
            issues.append(f"Cleaned {fixed} currency-formatted value(s) in {col}")

    # Convert to numeric
    df["Quantity"]    = pd.to_numeric(df["Quantity"],    errors="coerce")
    df["Unit Price"]  = pd.to_numeric(df["Unit Price"],  errors="coerce")
    df["Total Price"] = pd.to_numeric(df["Total Price"], errors="coerce")
    df["Discount %"]  = pd.to_numeric(df["Discount %"],  errors="coerce")

    # ── 10. Validate & fix negative quantities ───────────────────────────────
    neg_qty = df["Quantity"] < 0
    if neg_qty.any():
        count = neg_qty.sum()
        df.loc[neg_qty, "Quantity"] = df.loc[neg_qty, "Quantity"].abs()
        log(f"[Quantity]       Converted {count} negative quantity(ies) to absolute value")
        issues.append(f"Fixed {count} negative quantity value(s)")

    # ── 11. Round quantity to integer ────────────────────────────────────────
    float_qty = df["Quantity"].apply(lambda x: pd.notna(x) and x != int(x) if pd.notna(x) else False)
    if float_qty.any():
        count = float_qty.sum()
        df.loc[float_qty, "Quantity"] = df.loc[float_qty, "Quantity"].round().astype("Int64")
        log(f"[Quantity]       Rounded {count} float quantity value(s) to integer")
        issues.append(f"Rounded {count} float quantity(ies) to integer")

    # ── 12. Validate negative prices ─────────────────────────────────────────
    for col in ["Unit Price", "Total Price"]:
        neg = df[col] < 0
        if neg.any():
            count = neg.sum()
            df.loc[neg, col] = df.loc[neg, col].abs()
            log(f"[{col:<12}] Fixed {count} negative price(s) to absolute value")
            issues.append(f"Fixed {count} negative value(s) in {col}")

    # ── 13. Validate zero prices ─────────────────────────────────────────────
    zero_price = df["Unit Price"] == 0
    if zero_price.any():
        count = zero_price.sum()
        log(f"[Unit Price]     ⚠  {count} row(s) have a Unit Price of 0 — flagged for review")
        # Mark via a separate notes column instead of corrupting numeric dtype
        if "Notes" not in df.columns:
            df["Notes"] = ""
        df.loc[zero_price, "Notes"] = df.loc[zero_price, "Notes"].apply(
            lambda x: (x + "; " if x else "") + "REVIEW: Unit Price = 0"
        )
        issues.append(f"{count} row(s) have Unit Price = 0")

    # ── 14. Recalculate & validate Total Price ───────────────────────────────
    mismatch = 0
    for i, row in df.iterrows():
        try:
            qty   = float(row["Quantity"])
            price = float(row["Unit Price"])
            disc  = float(row["Discount %"]) if pd.notna(row["Discount %"]) else 0
            expected = round(qty * price * (1 - disc / 100), 2)
            actual   = float(row["Total Price"])
            if abs(expected - actual) > 0.02:
                df.at[i, "Total Price"] = expected
                mismatch += 1
        except (ValueError, TypeError):
            pass
    if mismatch:
        log(f"[Total Price]    Recalculated {mismatch} incorrect total(s)")
        issues.append(f"Recalculated {mismatch} wrong Total Price value(s)")

    # ── 15. Validate discount range ──────────────────────────────────────────
    bad_disc = df["Discount %"].apply(lambda x: pd.notna(x) and (x < 0 or x > 100))
    if bad_disc.any():
        count = bad_disc.sum()
        df.loc[bad_disc, "Discount %"] = None
        log(f"[Discount %]     Nulled {count} out-of-range discount(s) (>100% or <0%)")
        issues.append(f"Removed {count} invalid discount value(s)")

    # ── 16. Standardize Category ─────────────────────────────────────────────
    df["Category"] = df["Category"].str.title()
    invalid_cat = ~df["Category"].isin(VALID_CATEGORIES) & df["Category"].notna() & (df["Category"] != "")
    if invalid_cat.any():
        count = invalid_cat.sum()
        log(f"[Category]       ⚠  {count} unrecognized category value(s) flagged")
        if "Notes" not in df.columns:
            df["Notes"] = ""
        for i in df[invalid_cat].index:
            note = f"REVIEW: unknown category '{df.at[i, 'Category']}'"
            df.at[i, "Notes"] = (df.at[i, "Notes"] + "; " if df.at[i, "Notes"] else "") + note
        issues.append(f"{count} unrecognized category(ies)")
    else:
        log("[Category]       Standardized casing for all category values")

    # ── 17. Validate Status ──────────────────────────────────────────────────
    df["Status"] = df["Status"].str.strip().str.title()
    bad_status = ~df["Status"].isin(VALID_STATUSES) & df["Status"].notna() & (df["Status"] != "")
    if bad_status.any():
        count = bad_status.sum()
        if "Notes" not in df.columns:
            df["Notes"] = ""
        for i in df[bad_status].index:
            val = df.at[i, "Status"]
            fixed = False
            for valid in VALID_STATUSES:
                if val.lower().startswith(valid[:4].lower()):
                    df.at[i, "Status"] = valid
                    fixed = True
                    break
            if not fixed:
                note = f"REVIEW: unknown status '{val}'"
                df.at[i, "Notes"] = (df.at[i, "Notes"] + "; " if df.at[i, "Notes"] else "") + note
                df.at[i, "Status"] = "UNKNOWN"
        log(f"[Status]         Fixed/flagged {count} invalid status value(s)")
        issues.append(f"Fixed {count} invalid status value(s)")

    # ── 18. Standardize Region ───────────────────────────────────────────────
    df["Region"] = df["Region"].str.strip().str.title()
    missing_region = df["Region"].isna() | (df["Region"] == "")
    if missing_region.any():
        count = missing_region.sum()
        df.loc[missing_region, "Region"] = "UNKNOWN"
        log(f"[Region]         Filled {count} missing region(s) with 'UNKNOWN'")
        issues.append(f"Filled {count} missing region(s)")

    # ── 19. Fill missing Sales Rep ───────────────────────────────────────────
    missing_rep = df["Sales Rep"].isna() | (df["Sales Rep"] == "")
    if missing_rep.any():
        count = missing_rep.sum()
        df.loc[missing_rep, "Sales Rep"] = "UNASSIGNED"
        log(f"[Sales Rep]      Filled {count} missing sales rep(s) with 'UNASSIGNED'")
        issues.append(f"Filled {count} missing sales rep(s)")

    # ── 20. Add a data quality flag column ───────────────────────────────────
    def quality_flag(row):
        flags = []
        if "INVALID" in str(row.get("Email", "")):        flags.append("bad_email")
        if "REVIEW"  in str(row.get("Notes", "")):        flags.append("needs_review")
        if str(row.get("Order Date", "")) == "INVALID DATE": flags.append("bad_date")
        if str(row.get("Region", ""))    == "UNKNOWN":    flags.append("missing_region")
        if str(row.get("Sales Rep", "")) == "UNASSIGNED": flags.append("missing_rep")
        return ", ".join(flags) if flags else "OK"

    df["QA Flag"] = df.apply(quality_flag, axis=1)
    flagged = (df["QA Flag"] != "OK").sum()
    log(f"[QA Flag]        Added quality flag column — {flagged} row(s) need review")

    print("\n" + "─" * 60)
    print(f"\n  ✅  Cleaning complete!")
    print(f"      Rows processed  : {original_count}")
    print(f"      Rows after clean: {len(df)}")
    print(f"      Issues fixed    : {len(issues)}")
    print(f"      Rows flagged    : {flagged}")

    # ── Write cleaned Excel output ───────────────────────────────────────────
    save_clean_excel(df, output_path, issues)
    print(f"\n  📁  Saved to: {output_path}\n")


def save_clean_excel(df: pd.DataFrame, output_file: str, issues: list):
    """Write the cleaned dataframe to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Cleaned Data ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cleaned Data"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill   = PatternFill("solid", start_color="1F3864")
    ok_fill       = PatternFill("solid", start_color="E8F5E9")
    review_fill   = PatternFill("solid", start_color="FFF3CD")
    alt_fill      = PatternFill("solid", start_color="F5F8FF")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 22

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_review = any("REVIEW" in str(v) or "INVALID" in str(v) or "INVALID DATE" in str(v) for v in row)
        row_fill  = review_fill if is_review else (alt_fill if row_idx % 2 == 0 else None)
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    # ── Sheet 2: Issues Log ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues Log")
    ws2.append(["#", "Issue Description"])
    for cell in ws2[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", start_color="C0392B")
        cell.alignment = Alignment(horizontal="center")

    for i, issue in enumerate(issues, 1):
        ws2.append([i, issue])
        for cell in ws2[ws2.max_row]:
            cell.font   = Font(name="Arial", size=10)
            cell.border = border
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color="FDECEA")

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 60

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    summary_data = [
        ("Metric", "Value"),
        ("Total Rows (cleaned)",  len(df)),
        ("Rows needing review",   int((df["QA Flag"] != "OK").sum())),
        ("Total issues found",    len(issues)),
        ("Unique customers",      df["Customer Name"].nunique()),
        ("Unique products",       df["Product"].nunique()),
        ("Date range (min)",      df["Order Date"].min()),
        ("Date range (max)",      df["Order Date"].max()),
    ]
    for r_idx, (label, val) in enumerate(summary_data, 1):
        ws3.cell(r_idx, 1, label).font = Font(bold=(r_idx == 1), name="Arial")
        ws3.cell(r_idx, 2, val).font   = Font(name="Arial")
        if r_idx == 1:
            for c in [ws3.cell(r_idx, 1), ws3.cell(r_idx, 2)]:
                c.fill = PatternFill("solid", start_color="1F3864")
                c.font = Font(bold=True, color="FFFFFF", name="Arial")

    wb.save(output_file)


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Error: please provide an input file.")
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]

    if not Path(input_path).exists():
        print(f"Error: file not found: {input_path}")
        sys.exit(1)

    output_file = Path(input_path).stem + "_cleaned.xlsx"
    clean_sales_data(input_path, output_file)


if __name__ == "__main__":
    main()
