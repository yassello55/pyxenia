# ============================================================
#  ADVANCED WEB SCRAPER
#  Input  : websites_input_template.xlsx
#  Output : scraping_results.xlsx
# ============================================================
#
#  INSTALL DEPENDENCIES:
#    pip install requests beautifulsoup4 openpyxl pandas lxml fake-useragent
#
#  RUN:
#    python web_scraper.py
#    python web_scraper.py my_sites.xlsx          # custom input file
# ============================================================

import sys
import time
import random
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "websites_input_template.xlsx"
OUTPUT_FILE = "scraping_results.xlsx"
LOG_FILE    = "scraper.log"

HEADERS_POOL = [
    {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"},
    {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/605.1.15"},
    {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/119.0.0.0 Safari/537.36"},
]

REQUEST_TIMEOUT = 15       # seconds per request
DELAY_BETWEEN   = (1, 3)   # random wait between requests (seconds)
MAX_ITEMS       = 100      # max rows to scrape per site
MAX_RETRIES     = 2        # retry failed requests

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# STEP 1 — READ INPUT EXCEL
# ─────────────────────────────────────────────
def load_sites(path: str) -> list[dict]:
    """Read the Websites sheet and return a list of site configs."""
    try:
        df = pd.read_excel(path, sheet_name="Websites")
    except Exception as e:
        log.error(f"Cannot read input file: {e}")
        sys.exit(1)

    df.columns = [c.strip() for c in df.columns]
    required = {"URL", "Website Name", "Scrape Type"}
    if not required.issubset(df.columns):
        log.error(f"Missing columns in input file. Required: {required}")
        sys.exit(1)

    sites = []
    for _, row in df.iterrows():
        url = str(row.get("URL", "")).strip()
        if not url or url.lower() == "nan" or not url.startswith("http"):
            continue
        sites.append({
            "name":         str(row.get("Website Name", "Site")).strip(),
            "url":          url,
            "scrape_type":  str(row.get("Scrape Type", "custom")).strip().lower(),
            "tag":          str(row.get("Target Tag", "div")).strip(),
            "css_class":    str(row.get("Target Class", "")).strip(),
            "notes":        str(row.get("Notes", "")).strip(),
        })

    log.info(f"📋 Loaded {len(sites)} site(s) from {path}")
    return sites


# ─────────────────────────────────────────────
# STEP 2 — HTTP FETCH WITH RETRY
# ─────────────────────────────────────────────
def fetch(url: str) -> BeautifulSoup | None:
    """Download a page and return a BeautifulSoup object."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            headers = random.choice(HEADERS_POOL)
            resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except requests.exceptions.HTTPError as e:
            log.warning(f"  HTTP {e.response.status_code} on attempt {attempt}: {url}")
        except requests.exceptions.ConnectionError:
            log.warning(f"  Connection error on attempt {attempt}: {url}")
        except requests.exceptions.Timeout:
            log.warning(f"  Timeout on attempt {attempt}: {url}")
        except Exception as e:
            log.warning(f"  Unexpected error on attempt {attempt}: {e}")

        if attempt < MAX_RETRIES:
            time.sleep(2)

    log.error(f"  ❌ Failed after {MAX_RETRIES} attempts: {url}")
    return None


# ─────────────────────────────────────────────
# STEP 3 — SCRAPING STRATEGIES
# ─────────────────────────────────────────────
def scrape_product_list(soup: BeautifulSoup, tag: str, css_class: str) -> list[dict]:
    """Extract product name + price."""
    items = soup.find_all(tag, class_=css_class or True)[:MAX_ITEMS]
    results = []
    for item in items:
        title = (item.find("h3") or item.find("h2") or item.find("a"))
        price = item.find(class_=lambda c: c and "price" in c.lower()) if item else None
        results.append({
            "Title":       title.get_text(strip=True) if title else "",
            "Price":       price.get_text(strip=True) if price else "",
            "Link":        title.find("a")["href"] if title and title.find("a") else "",
        })
    return results


def scrape_article_list(soup: BeautifulSoup, tag: str, css_class: str) -> list[dict]:
    """Extract article headlines + links."""
    items = soup.find_all(tag, class_=css_class or True)[:MAX_ITEMS]
    results = []
    for item in items:
        a = item.find("a") or item
        results.append({
            "Headline": a.get_text(strip=True),
            "Link":     a.get("href", ""),
        })
    return [r for r in results if r["Headline"]]


def scrape_quotes(soup: BeautifulSoup, tag: str, css_class: str) -> list[dict]:
    """Extract quote text + author."""
    items = soup.find_all(tag, class_=css_class or True)[:MAX_ITEMS]
    results = []
    for item in items:
        text   = item.find(class_="text")
        author = item.find(class_="author")
        tags   = [t.get_text(strip=True) for t in item.find_all(class_="tag")]
        results.append({
            "Quote":  text.get_text(strip=True)   if text   else "",
            "Author": author.get_text(strip=True) if author else "",
            "Tags":   ", ".join(tags),
        })
    return [r for r in results if r["Quote"]]


def scrape_custom(soup: BeautifulSoup, tag: str, css_class: str) -> list[dict]:
    """Extract any text content from matching elements."""
    find_args = {"class_": css_class} if css_class else {}
    items = soup.find_all(tag, **find_args)[:MAX_ITEMS]
    return [{"Text": item.get_text(strip=True), "HTML Tag": item.name}
            for item in items if item.get_text(strip=True)]


STRATEGIES = {
    "product_list": scrape_product_list,
    "article_list": scrape_article_list,
    "quotes":       scrape_quotes,
    "custom":       scrape_custom,
}


# ─────────────────────────────────────────────
# STEP 4 — SCRAPE ONE SITE
# ─────────────────────────────────────────────
def scrape_site(site: dict) -> dict:
    """Scrape a single site and return structured result."""
    log.info(f"🌐 Scraping: {site['name']} ({site['url']})")
    result = {
        "name":    site["name"],
        "url":     site["url"],
        "status":  "pending",
        "rows":    [],
        "error":   "",
        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    soup = fetch(site["url"])
    if soup is None:
        result["status"] = "failed"
        result["error"]  = "Could not connect to site"
        return result

    strategy = STRATEGIES.get(site["scrape_type"], scrape_custom)
    try:
        rows = strategy(soup, site["tag"], site["css_class"])
        result["rows"]   = rows
        result["status"] = "success" if rows else "empty"
        log.info(f"  ✅ {len(rows)} items found")
    except Exception as e:
        result["status"] = "error"
        result["error"]  = str(e)
        log.error(f"  ❌ Scraping error: {e}")

    # Polite delay
    delay = random.uniform(*DELAY_BETWEEN)
    time.sleep(delay)

    return result


# ─────────────────────────────────────────────
# STEP 5 — WRITE OUTPUT EXCEL
# ─────────────────────────────────────────────
DARK_BLUE   = "2F4F8F"
GREEN       = "1A7A4A"
RED         = "C0392B"
LIGHT_GREY  = "F5F5F5"
WHITE       = "FFFFFF"

def styled_header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def write_data_row(ws, row_i, values, is_alt):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", start_color=LIGHT_GREY) if is_alt else None
    for col_i, val in enumerate(values, 1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def write_results_to_excel(results: list[dict], output_path: str):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ──────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 42
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 14
    ws_sum.column_dimensions["F"].width = 20
    ws_sum.column_dimensions["G"].width = 28

    sum_headers = ["#", "Website Name", "URL", "Status", "Items Found", "Scraped At", "Notes / Error"]
    for col_i, h in enumerate(sum_headers, 1):
        styled_header_cell(ws_sum, 1, col_i, h)
    ws_sum.row_dimensions[1].height = 28

    for i, r in enumerate(results, 1):
        status_color = GREEN if r["status"] == "success" else RED
        row_vals = [
            i,
            r["name"],
            r["url"],
            r["status"].upper(),
            len(r["rows"]),
            r["scraped_at"],
            r["error"] or "",
        ]
        write_data_row(ws_sum, i + 1, row_vals, i % 2 == 0)

        # Colour the status cell
        status_cell = ws_sum.cell(row=i + 1, column=4)
        status_cell.font = Font(name="Arial", size=10, bold=True, color=status_color)

    # ── One sheet per site ─────────────────────────────────
    for r in results:
        if not r["rows"]:
            continue

        # Safe sheet name (max 31 chars, no special chars)
        sheet_name = r["name"][:28].replace("/", "-").replace(":", "").replace("?", "")
        ws = wb.create_sheet(sheet_name)

        # Meta row
        ws.merge_cells("A1:G1")
        meta = ws["A1"]
        meta.value = f"Source: {r['url']}  |  Scraped: {r['scraped_at']}  |  Items: {len(r['rows'])}"
        meta.font  = Font(name="Arial", size=9, italic=True, color="666666")
        meta.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20

        # Column headers from first row keys
        col_keys = list(r["rows"][0].keys()) + ["Source URL", "Scraped At"]
        for col_i, key in enumerate(col_keys, 1):
            styled_header_cell(ws, 2, col_i, key)
            ws.column_dimensions[get_column_letter(col_i)].width = max(18, len(key) + 4)
        ws.row_dimensions[2].height = 26

        # Data rows
        for row_i, item in enumerate(r["rows"], 3):
            vals = [item.get(k, "") for k in list(r["rows"][0].keys())]
            vals += [r["url"], r["scraped_at"]]
            write_data_row(ws, row_i, vals, row_i % 2 == 0)

        ws.freeze_panes = "A3"

    # ── Save ───────────────────────────────────────────────
    wb.save(output_path)
    log.info(f"\n💾 Results saved to: {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    log.info("=" * 55)
    log.info("  WEB SCRAPER STARTED")
    log.info(f"  Input  : {INPUT_FILE}")
    log.info(f"  Output : {OUTPUT_FILE}")
    log.info("=" * 55)

    sites   = load_sites(INPUT_FILE)
    results = [scrape_site(site) for site in sites]

    success = sum(1 for r in results if r["status"] == "success")
    failed  = sum(1 for r in results if r["status"] in ("failed", "error"))
    total   = sum(len(r["rows"]) for r in results)

    write_results_to_excel(results, OUTPUT_FILE)

    log.info("\n── FINAL SUMMARY ──────────────────────────────")
    log.info(f"  Sites processed : {len(sites)}")
    log.info(f"  Successful      : {success}")
    log.info(f"  Failed          : {failed}")
    log.info(f"  Total items     : {total}")
    log.info(f"  Output file     : {OUTPUT_FILE}")
    log.info("───────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
